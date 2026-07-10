import io
import re
from collections import defaultdict
from datetime import date, datetime, timedelta

import pandas as pd
import streamlit as st

import utils


# =========================================================
# ΜΗΝΙΑΙΕΣ ΩΡΕΣ ΕΡΓΩΝ
#
# Ανεξάρτητη σελίδα που εμφανίζεται αυτόματα στο Streamlit
# sidebar επειδή βρίσκεται μέσα στον φάκελο pages.
#
# Δεν αλλάζει assignments, έργα ή εργαζομένους.
# Διαβάζει μόνο τα υπάρχοντα δεδομένα και δημιουργεί αναφορά.
# =========================================================

st.set_page_config(
    page_title="Μηνιαίες Ώρες Έργων",
    page_icon="🕒",
    layout="wide",
)


# --- AUTH / SESSION SAFETY ---
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.get("authenticated"):
    st.switch_page("streamlit_app.py")
    st.stop()

for key, default in {
    "employees": [],
    "projects": [],
    "assignments": [],
    "leaves": [],
    "recurring_patterns": [],
    "evaluations": [],
}.items():
    if key not in st.session_state:
        st.session_state[key] = default


# --- DATA / COMMON SIDEBAR ---
utils.init_data_and_sync()
utils.setup_shared_ui()


GREEK_MONTHS = {
    1: "Ιανουάριος",
    2: "Φεβρουάριος",
    3: "Μάρτιος",
    4: "Απρίλιος",
    5: "Μάιος",
    6: "Ιούνιος",
    7: "Ιούλιος",
    8: "Αύγουστος",
    9: "Σεπτέμβριος",
    10: "Οκτώβριος",
    11: "Νοέμβριος",
    12: "Δεκέμβριος",
}

GREEK_DAYS = {
    0: "Δευτέρα",
    1: "Τρίτη",
    2: "Τετάρτη",
    3: "Πέμπτη",
    4: "Παρασκευή",
    5: "Σάββατο",
    6: "Κυριακή",
}


def _safe_date(value):
    if isinstance(value, date) and not isinstance(value, datetime):
        return value

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, str):
        raw = value.split("T")[0][:10]
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue

    return None


def _safe_time_text(value):
    raw = str(value or "").strip()
    if not raw:
        return ""
    return raw[:5]


def _time_to_minutes(value):
    raw = _safe_time_text(value)
    if not raw or ":" not in raw:
        return None

    try:
        hour, minute = map(int, raw.split(":"))
    except Exception:
        return None

    return hour * 60 + minute


def _duration_hours(start_value, end_value):
    start_minutes = _time_to_minutes(start_value)
    end_minutes = _time_to_minutes(end_value)

    if start_minutes is None or end_minutes is None:
        return 0.0

    if end_minutes <= start_minutes:
        end_minutes += 24 * 60

    return round((end_minutes - start_minutes) / 60.0, 2)


def _format_hours(value):
    value = float(value or 0)

    if abs(value - round(value)) < 0.001:
        return str(int(round(value)))

    return f"{value:.2f}".rstrip("0").rstrip(".")


def _project_map():
    return {
        project.get("id"): project
        for project in st.session_state.get("projects", [])
        if isinstance(project, dict) and project.get("id")
    }


def _employee_map():
    return {
        employee.get("id"): employee
        for employee in st.session_state.get("employees", [])
        if isinstance(employee, dict) and employee.get("id")
    }


def _project_name(project_id):
    project = _project_map().get(project_id)
    return project.get("name", "Άγνωστο Έργο") if project else "Άγνωστο Έργο"


def _employee_name(employee_id):
    employee = _employee_map().get(employee_id)
    return employee.get("name", "Άγνωστος") if employee else "Άγνωστος"


def _employee_phone(employee_id):
    employee = _employee_map().get(employee_id)
    return str(employee.get("phone", "") or "") if employee else ""


def _month_bounds(year, month):
    first_day = date(year, month, 1)

    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)

    return first_day, next_month - timedelta(days=1)


def _available_years():
    years = set()

    for assignment in st.session_state.get("assignments", []):
        if not isinstance(assignment, dict):
            continue

        work_date = _safe_date(assignment.get("date"))
        if work_date:
            years.add(work_date.year)

    current_year = date.today().year
    years.add(current_year)

    return sorted(years, reverse=True)


def _month_assignments(year, month, selected_project_ids):
    month_start, month_end = _month_bounds(year, month)
    selected_project_ids = set(selected_project_ids or [])

    rows = []

    for assignment in st.session_state.get("assignments", []):
        if not isinstance(assignment, dict):
            continue

        if assignment.get("is_cancelled", False):
            continue

        employee_id = assignment.get("employeeId")
        project_id = assignment.get("projectId")

        # Δεν υπολογίζουμε εγγραφές χωρίς εργαζόμενο.
        if not employee_id or not project_id:
            continue

        if selected_project_ids and project_id not in selected_project_ids:
            continue

        work_date = _safe_date(assignment.get("date"))
        if not work_date or not (month_start <= work_date <= month_end):
            continue

        start_time = _safe_time_text(assignment.get("startTime"))
        end_time = _safe_time_text(assignment.get("endTime"))
        hours = _duration_hours(start_time, end_time)

        if hours <= 0:
            continue

        rows.append(
            {
                "assignment_id": assignment.get("id"),
                "date": work_date,
                "day_name": GREEK_DAYS.get(work_date.weekday(), ""),
                "is_sunday": work_date.weekday() == 6,
                "project_id": project_id,
                "project": _project_name(project_id),
                "employee_id": employee_id,
                "employee": _employee_name(employee_id),
                "phone": _employee_phone(employee_id),
                "start": start_time,
                "end": end_time,
                "schedule": f"{start_time} - {end_time}",
                "hours": hours,
                "notes": str(assignment.get("notes", "") or ""),
                "recurring_id": assignment.get("recurring_id"),
            }
        )

    rows.sort(
        key=lambda row: (
            row["project"].lower(),
            row["employee"].lower(),
            row["date"],
            row["start"],
        )
    )

    return rows


def _combine_same_day_shifts(rows):
    """
    Συνδυάζει πολλαπλές βάρδιες ίδιου εργαζομένου, έργου και ημέρας.

    Παράδειγμα:
    07:00-11:00 και 16:00-18:00
    εμφανίζονται στο ίδιο κελί ως:
    07:00 - 11:00 / 16:00 - 18:00
    """
    grouped = defaultdict(list)

    for row in rows:
        key = (row["project_id"], row["employee_id"], row["date"])
        grouped[key].append(row)

    result = {}

    for key, shifts in grouped.items():
        shifts = sorted(shifts, key=lambda item: item["start"])
        schedule = " / ".join(shift["schedule"] for shift in shifts)
        total_hours = round(sum(shift["hours"] for shift in shifts), 2)

        result[key] = {
            "schedule": schedule,
            "hours": total_hours,
            "is_sunday": shifts[0]["is_sunday"],
            "notes": " | ".join(
                dict.fromkeys(
                    shift["notes"].strip()
                    for shift in shifts
                    if shift["notes"].strip()
                )
            ),
        }

    return result


def _summary_dataframe(rows):
    columns = [
        "Έργο",
        "Εργαζόμενος",
        "Τηλέφωνο",
        "Ημέρες Εργασίας",
        "Βάρδιες",
        "Ώρες Καθημερινών",
        "Ώρες Κυριακών",
        "Σύνολο Ωρών",
    ]

    if not rows:
        return pd.DataFrame(columns=columns)

    df = pd.DataFrame(rows)

    summary = (
        df.groupby(
            ["project_id", "project", "employee_id", "employee", "phone"],
            as_index=False,
        )
        .agg(
            work_days=("date", "nunique"),
            shift_count=("assignment_id", "count"),
            weekday_hours=(
                "hours",
                lambda values: round(
                    sum(
                        value
                        for value, is_sunday in zip(
                            values,
                            df.loc[values.index, "is_sunday"],
                        )
                        if not is_sunday
                    ),
                    2,
                ),
            ),
            sunday_hours=(
                "hours",
                lambda values: round(
                    sum(
                        value
                        for value, is_sunday in zip(
                            values,
                            df.loc[values.index, "is_sunday"],
                        )
                        if is_sunday
                    ),
                    2,
                ),
            ),
            total_hours=("hours", "sum"),
        )
        .sort_values(["project", "employee"])
    )

    summary["total_hours"] = summary["total_hours"].round(2)

    summary = summary.rename(
        columns={
            "project": "Έργο",
            "employee": "Εργαζόμενος",
            "phone": "Τηλέφωνο",
            "work_days": "Ημέρες Εργασίας",
            "shift_count": "Βάρδιες",
            "weekday_hours": "Ώρες Καθημερινών",
            "sunday_hours": "Ώρες Κυριακών",
            "total_hours": "Σύνολο Ωρών",
        }
    )

    return summary[columns]


def _detail_dataframe(rows):
    columns = [
        "Ημερομηνία",
        "Ημέρα",
        "Έργο",
        "Εργαζόμενος",
        "Τηλέφωνο",
        "Έναρξη",
        "Λήξη",
        "Ώρες",
        "Τύπος Ώρας",
        "Παρατηρήσεις",
    ]

    detail_rows = []

    for row in rows:
        detail_rows.append(
            {
                "Ημερομηνία": row["date"].strftime("%d/%m/%Y"),
                "Ημέρα": row["day_name"],
                "Έργο": row["project"],
                "Εργαζόμενος": row["employee"],
                "Τηλέφωνο": row["phone"],
                "Έναρξη": row["start"],
                "Λήξη": row["end"],
                "Ώρες": row["hours"],
                "Τύπος Ώρας": "Κυριακή" if row["is_sunday"] else "Καθημερινή",
                "Παρατηρήσεις": row["notes"],
            }
        )

    return pd.DataFrame(detail_rows, columns=columns)


def _safe_sheet_title(value, existing_titles):
    title = re.sub(r'[\\/*?:\[\]]', "-", str(value or "Έργο")).strip()
    title = title[:31] or "Έργο"

    base = title
    counter = 2

    while title in existing_titles:
        suffix = f" ({counter})"
        title = f"{base[:31-len(suffix)]}{suffix}"
        counter += 1

    existing_titles.add(title)
    return title


def _style_report_sheet(ws, last_row, last_col):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    dark_green = "4F7F35"
    medium_green = "93C47D"
    pale_green = "D9EAD3"
    pale_yellow = "FFE599"
    orange = "F9B233"
    dark_text = "1F1F1F"

    thin_green = Side(style="thin", color="93C47D")
    medium_dark_green = Side(style="medium", color="274E13")

    grid_border = Border(
        left=thin_green,
        right=thin_green,
        top=thin_green,
        bottom=thin_green,
    )

    header_border = Border(
        left=medium_dark_green,
        right=medium_dark_green,
        top=medium_dark_green,
        bottom=medium_dark_green,
    )

    # Title
    ws["A1"].fill = PatternFill("solid", fgColor=dark_green)
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = header_border

    # Employee group headers
    for col in range(1, last_col + 1):
        cell = ws.cell(row=2, column=col)
        cell.fill = PatternFill("solid", fgColor=medium_green)
        cell.font = Font(color=dark_text, bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border

    # Column headers
    for col in range(1, last_col + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = PatternFill("solid", fgColor=orange)
        cell.font = Font(color=dark_text, bold=True, size=8)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border

    # Body
    for row in range(4, last_row + 1):
        date_cell = ws.cell(row=row, column=1)
        date_cell.fill = PatternFill("solid", fgColor=pale_yellow)
        date_cell.font = Font(bold=True, size=8)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        date_cell.border = grid_border

        for col in range(2, last_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=pale_green)
            cell.font = Font(size=8)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = grid_border

    # Widths
    ws.column_dimensions["A"].width = 14

    for col in range(2, last_col + 1):
        mod = (col - 2) % 3

        if mod == 0:
            ws.column_dimensions[get_column_letter(col)].width = 22
        else:
            ws.column_dimensions[get_column_letter(col)].width = 13

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 30

    for row in range(4, last_row + 1):
        ws.row_dimensions[row].height = 23

    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _add_summary_sheet(wb, summary_df, year, month):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Σύνολα", 0)

    title = f"Συγκεντρωτικές Ώρες — {GREEK_MONTHS[month]} {year}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=1, column=1).value = title

    headers = list(summary_df.columns)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=3, column=col_idx).value = header

    for row_idx, row in enumerate(summary_df.itertuples(index=False), start=4):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    dark_green = "4F7F35"
    orange = "F9B233"
    pale_green = "EAF4E3"
    thin_green = Side(style="thin", color="93C47D")

    border = Border(
        left=thin_green,
        right=thin_green,
        top=thin_green,
        bottom=thin_green,
    )

    ws["A1"].fill = PatternFill("solid", fgColor=dark_green)
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, 9):
        cell = ws.cell(row=3, column=col)
        cell.fill = PatternFill("solid", fgColor=orange)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in range(4, 4 + len(summary_df)):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill("solid", fgColor=pale_green)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [25, 28, 16, 16, 10, 18, 16, 14]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


def _build_monthly_excel(rows, year, month):
    from openpyxl import Workbook

    workbook = Workbook()
    workbook.remove(workbook.active)

    combined = _combine_same_day_shifts(rows)
    summary_df = _summary_dataframe(rows)
    _add_summary_sheet(workbook, summary_df, year, month)

    month_start, month_end = _month_bounds(year, month)
    all_dates = []
    cursor = month_start

    while cursor <= month_end:
        all_dates.append(cursor)
        cursor += timedelta(days=1)

    projects = sorted(
        {
            (row["project_id"], row["project"])
            for row in rows
        },
        key=lambda item: item[1].lower(),
    )

    existing_titles = {"Σύνολα"}

    for project_id, project_name in projects:
        project_rows = [
            row for row in rows
            if row["project_id"] == project_id
        ]

        employee_ids = sorted(
            {row["employee_id"] for row in project_rows},
            key=lambda employee_id: _employee_name(employee_id).lower(),
        )

        sheet_title = _safe_sheet_title(project_name, existing_titles)
        ws = workbook.create_sheet(sheet_title)

        last_col = 1 + len(employee_ids) * 3

        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=max(1, last_col),
        )
        ws.cell(row=1, column=1).value = (
            f"{project_name} — {GREEK_MONTHS[month]} {year}"
        )

        ws.cell(row=2, column=1).value = project_name
        ws.cell(row=3, column=1).value = "ΗΜΕΡΕΣ"

        for employee_index, employee_id in enumerate(employee_ids):
            start_col = 2 + employee_index * 3
            end_col = start_col + 2

            employee_label = _employee_name(employee_id)
            phone = _employee_phone(employee_id)

            if phone:
                employee_label = f"{employee_label}\nτηλ. {phone}"

            ws.merge_cells(
                start_row=2,
                start_column=start_col,
                end_row=2,
                end_column=end_col,
            )
            ws.cell(row=2, column=start_col).value = employee_label

            ws.cell(row=3, column=start_col).value = "ώρες"
            ws.cell(row=3, column=start_col + 1).value = "σύνολο ωρών καθημερινές"
            ws.cell(row=3, column=start_col + 2).value = "ΚΥΡΙΑΚΗ"

        for row_offset, current_date in enumerate(all_dates, start=4):
            ws.cell(row=row_offset, column=1).value = current_date
            ws.cell(row=row_offset, column=1).number_format = "dd/mm/yyyy"

            for employee_index, employee_id in enumerate(employee_ids):
                start_col = 2 + employee_index * 3
                key = (project_id, employee_id, current_date)
                day_data = combined.get(key)

                if not day_data:
                    continue

                ws.cell(row=row_offset, column=start_col).value = day_data["schedule"]

                if day_data["is_sunday"]:
                    ws.cell(row=row_offset, column=start_col + 2).value = day_data["hours"]
                else:
                    ws.cell(row=row_offset, column=start_col + 1).value = day_data["hours"]

        total_row = 4 + len(all_dates) + 1
        ws.cell(row=total_row, column=1).value = "ΣΥΝΟΛΑ"

        for employee_index, employee_id in enumerate(employee_ids):
            start_col = 2 + employee_index * 3
            weekday_col = start_col + 1
            sunday_col = start_col + 2

            weekday_total = sum(
                row["hours"]
                for row in project_rows
                if row["employee_id"] == employee_id and not row["is_sunday"]
            )
            sunday_total = sum(
                row["hours"]
                for row in project_rows
                if row["employee_id"] == employee_id and row["is_sunday"]
            )

            ws.cell(row=total_row, column=weekday_col).value = round(weekday_total, 2)
            ws.cell(row=total_row, column=sunday_col).value = round(sunday_total, 2)

        _style_report_sheet(
            ws,
            last_row=total_row,
            last_col=max(1, last_col),
        )

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# --- PAGE UI ---
st.title("🕒 Μηνιαίες Ώρες Έργων")
st.caption(
    "Μηνιαία συγκεντρωτική αναφορά ωρών ανά έργο και εργαζόμενο. "
    "Η σελίδα διαβάζει μόνο τις καταχωρημένες, μη ακυρωμένες βάρδιες."
)

projects = sorted(
    [
        project
        for project in st.session_state.get("projects", [])
        if isinstance(project, dict) and project.get("id")
    ],
    key=lambda project: str(project.get("name", "")).lower(),
)

if not projects:
    st.warning("Δεν υπάρχουν έργα για εμφάνιση.")
    st.stop()

years = _available_years()
today = date.today()

col_year, col_month = st.columns(2)

with col_year:
    selected_year = st.selectbox(
        "Έτος",
        options=years,
        index=years.index(today.year) if today.year in years else 0,
    )

with col_month:
    selected_month = st.selectbox(
        "Μήνας",
        options=list(range(1, 13)),
        index=today.month - 1,
        format_func=lambda month: GREEK_MONTHS[month],
    )

project_options = [project["id"] for project in projects]

selected_project_ids = st.multiselect(
    "Έργα",
    options=project_options,
    default=project_options,
    format_func=_project_name,
    help="Άφησε επιλεγμένα όλα τα έργα για συνολική μηνιαία αναφορά.",
)

rows = _month_assignments(
    selected_year,
    selected_month,
    selected_project_ids,
)

summary_df = _summary_dataframe(rows)
detail_df = _detail_dataframe(rows)

month_start, month_end = _month_bounds(selected_year, selected_month)

st.caption(
    f"Περίοδος: {month_start.strftime('%d/%m/%Y')} – "
    f"{month_end.strftime('%d/%m/%Y')}"
)

if not rows:
    st.info("Δεν βρέθηκαν καταχωρημένες ώρες για τα επιλεγμένα φίλτρα.")
    st.stop()

total_hours = round(sum(row["hours"] for row in rows), 2)
weekday_hours = round(
    sum(row["hours"] for row in rows if not row["is_sunday"]),
    2,
)
sunday_hours = round(
    sum(row["hours"] for row in rows if row["is_sunday"]),
    2,
)
employee_count = len({row["employee_id"] for row in rows})

metric_1, metric_2, metric_3, metric_4 = st.columns(4)

metric_1.metric("Σύνολο ωρών", _format_hours(total_hours))
metric_2.metric("Καθημερινές", _format_hours(weekday_hours))
metric_3.metric("Κυριακές", _format_hours(sunday_hours))
metric_4.metric("Εργαζόμενοι", employee_count)

st.divider()

tab_summary, tab_detail = st.tabs(
    ["📊 Συγκεντρωτικά", "📋 Αναλυτικές βάρδιες"]
)

with tab_summary:
    st.dataframe(
        summary_df,
        use_container_width=True,
        hide_index=True,
    )

with tab_detail:
    st.dataframe(
        detail_df,
        use_container_width=True,
        hide_index=True,
    )

st.divider()

try:
    excel_bytes = _build_monthly_excel(
        rows,
        selected_year,
        selected_month,
    )

    st.download_button(
        label="📥 Εξαγωγή Μηνιαίων Ωρών σε Excel",
        data=excel_bytes,
        file_name=(
            f"Μηνιαίες_Ώρες_Έργων_"
            f"{selected_year}_{selected_month:02d}.xlsx"
        ),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
except Exception as exc:
    st.warning(
        f"Δεν ήταν δυνατή η δημιουργία του Excel: {exc}"
    )

st.info(
    "Η αναφορά υπολογίζει πραγματικές ώρες από τις μπάρες του Gantt. "
    "Οι εγγραφές χωρίς προσωπικό και οι ακυρωμένες βάρδιες δεν υπολογίζονται."
)
