import io
import math
from datetime import datetime, timedelta

import streamlit as st

import config


def _normalize_hex(color_value, default="F6B26B"):
    raw = str(color_value or default).strip().replace("#", "")
    if len(raw) == 3:
        raw = "".join(ch * 2 for ch in raw)
    if len(raw) != 6:
        raw = default
    return raw.upper()


def _font_color_for_fill(hex_color):
    """Επιστρέφει μαύρο ή λευκό ανάλογα με τη φωτεινότητα του φόντου."""
    raw = _normalize_hex(hex_color)
    r = int(raw[0:2], 16)
    g = int(raw[2:4], 16)
    b = int(raw[4:6], 16)
    luminance = (0.299 * r + 0.587 * g + 0.114 * b)
    return "000000" if luminance > 150 else "FFFFFF"


def _time_to_minutes(time_value):
    raw = str(time_value or "00:00")[:5]
    hour, minute = map(int, raw.split(":"))
    if hour < 4:
        hour += 24
    return (hour - 4) * 60 + minute


def _slot_index(time_value, slot_minutes=30, mode="floor"):
    minutes = _time_to_minutes(time_value)
    if mode == "ceil":
        return int(math.ceil(minutes / slot_minutes))
    return int(math.floor(minutes / slot_minutes))


def _shorten_text(value, max_len=115):
    text = str(value or "")
    if len(text) <= max_len:
        return text
    return text[: max_len - 1] + "…"


def _safe_employee_short_names():
    emp_short_names = {}
    for emp in st.session_state.get("employees", []):
        if not isinstance(emp, dict):
            continue
        emp_id = emp.get("id")
        full_name = emp.get("name", "")
        if not emp_id:
            continue
        parts = full_name.split()
        emp_short_names[emp_id] = f"{parts[-1]} {parts[0][0]}." if len(parts) > 1 else full_name
    return emp_short_names


def _employee_name(employee_id, emp_short_names=None):
    if not employee_id:
        return ""
    emp_short_names = emp_short_names or {}
    if employee_id in emp_short_names:
        return emp_short_names[employee_id]

    try:
        import utils
        return utils.get_employee_name(employee_id)
    except Exception:
        return str(employee_id)


def _day_left_label(current_date):
    """
    Κείμενο πληροφοριών ημέρας για το visual Excel Gantt.

    Περιλαμβάνει:
    - Άδειες ημέρας
    - Αντικαταστάτες
    - Διαθέσιμα εξωτερικά συνεργεία μετά τα πρωινά
    """
    emp_short_names = _safe_employee_short_names()
    lines = []

    leaves_today = []
    for leave in st.session_state.get("leaves", []):
        if not isinstance(leave, dict):
            continue

        start_date = leave.get("startDate")
        end_date = leave.get("endDate")
        if not start_date or not end_date:
            continue

        try:
            if start_date <= current_date <= end_date:
                emp_n = _employee_name(leave.get("employeeId"), emp_short_names)
                sub_id = leave.get("substituteId")
                if sub_id:
                    sub_n = _employee_name(sub_id, emp_short_names)
                    leaves_today.append(f"{emp_n} (Αντ: {sub_n})")
                elif emp_n:
                    leaves_today.append(emp_n)
        except Exception:
            continue

    if leaves_today:
        lines.append("Άδειες:")
        lines.extend(leaves_today)

    available_ext_crew = []
    day_assigns = st.session_state.get("assignments_by_date", {}).get(current_date, [])

    for emp in st.session_state.get("employees", []):
        if not isinstance(emp, dict):
            continue

        emp_id = emp.get("id")
        if not emp_id:
            continue

        if emp.get("status", "Ενεργός") != "Ενεργός":
            continue

        if not emp.get("is_external_crew", False):
            continue

        is_on_leave = False
        for leave in st.session_state.get("leaves", []):
            if not isinstance(leave, dict):
                continue
            try:
                if leave.get("employeeId") == emp_id and leave.get("startDate") <= current_date <= leave.get("endDate"):
                    is_on_leave = True
                    break
            except Exception:
                continue

        if is_on_leave:
            continue

        is_busy_after_10 = any(
            isinstance(a, dict)
            and a.get("employeeId") == emp_id
            and not a.get("is_cancelled", False)
            and str(a.get("endTime", ""))[:5] > "10:00"
            for a in day_assigns
        )

        if not is_busy_after_10:
            available_ext_crew.append(emp_short_names.get(emp_id, emp.get("name", "")))

    if available_ext_crew:
        if lines:
            lines.append("")
        lines.append("ΜΕΤΑ ΤΑ ΠΡΩΙΝΑ:")
        lines.extend(available_ext_crew)

    if not lines:
        lines.append("")

    return "\n".join(lines), len(lines)


def _place_groups_in_lanes(groups, lane_offset=0):
    """
    Τοποθετεί μπάρες σε lanes χωρίς επικάλυψη.
    Επιστρέφει λίστα (group, τελικό_lane_idx) και πόσα lanes χρησιμοποίησε.
    """
    lanes = []
    placed = []

    for group in sorted(groups, key=lambda x: (x.get("StartTime", ""), x.get("EndTime", ""), x.get("Project", ""))):
        start = str(group.get("StartTime", ""))[:5]
        end = str(group.get("EndTime", ""))[:5]

        placed_lane = None
        for lane_idx, lane_end in enumerate(lanes):
            if start >= lane_end:
                lanes[lane_idx] = end
                placed_lane = lane_idx
                break

        if placed_lane is None:
            lanes.append(end)
            placed_lane = len(lanes) - 1

        placed.append((group, lane_offset + placed_lane))

    return placed, len(lanes)


def _is_blue_stack_group(group):
    """
    True για τις μπλε μπάρες που στο HTML Gantt μπαίνουν πάντα στο κάτω μέρος της ημέρας.
    Χρησιμοποιεί το ίδιο BLUE_STACK_HEX από το config.
    """
    blue_stack_hex = str(getattr(config, "BLUE_STACK_HEX", "") or "").lower()
    group_color = str(group.get("ColorHex", "") or "").lower()
    return bool(blue_stack_hex and group_color == blue_stack_hex)


def _build_lanes(day_groups):
    """
    Ίδια λογική στοίβαξης με το HTML Gantt:
    - πρώτα τοποθετούνται όλες οι μη μπλε μπάρες
    - μετά τοποθετούνται οι μπλε μπάρες
    Έτσι οι μπλε μπάρες εμφανίζονται πάντα στο κάτω μέρος της ημέρας και στο Excel.
    """
    non_blue_groups = [group for group in (day_groups or []) if not _is_blue_stack_group(group)]
    blue_groups = [group for group in (day_groups or []) if _is_blue_stack_group(group)]

    non_blue_placed, non_blue_lane_count = _place_groups_in_lanes(non_blue_groups, lane_offset=0)
    blue_placed, blue_lane_count = _place_groups_in_lanes(blue_groups, lane_offset=non_blue_lane_count)

    placed = non_blue_placed + blue_placed
    total_lanes = non_blue_lane_count + blue_lane_count

    return placed, max(1, total_lanes)


def _apply_range_border(ws, min_row, max_row, min_col, max_col, border):
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).border = border


def create_visual_gantt_excel(wk_groups, start_of_week, slot_minutes=30):
    """
    Δημιουργεί Excel τύπου Gantt με συγχωνευμένα κελιά.

    Παλέτα/ύφος βασισμένο στο δείγμα Google Sheet:
    - ανοιχτό πράσινο φόντο grid
    - σκούρο πράσινο κελί ημέρας
    - πράσινη ημερομηνία
    - κίτρινη/πράσινη περιοχή πληροφοριών
    - πορτοκαλί κεφαλίδες ωρών
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise RuntimeError("Λείπει το openpyxl από το περιβάλλον.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Gantt"

    # Χρονικό εύρος όπως το HTML Gantt: 04:00 έως 00:00.
    start_hour = 4
    total_minutes = 20 * 60
    slot_count = total_minutes // slot_minutes

    # Στήλες A-D είναι οι πληροφορίες ημέρας. Οι ώρες ξεκινούν από E.
    day_col = 1
    date_col = 2
    info_col = 3
    spacer_col = 4
    first_time_col = 5
    last_time_col = first_time_col + slot_count - 1

    # --- Παλέτα από το screenshot ---
    dark_green = "4F7F35"
    medium_green = "93C47D"
    light_green = "D9EAD3"
    very_light_green = "EAF4E3"
    header_orange = "F9B233"
    bar_orange = "F6B26B"
    pale_yellow = "FFE599"
    bright_green = "00F000"
    purple_separator = "4C1130"
    dark_text = "1F1F1F"
    red = "DC2626"

    # Fills
    dark_green_fill = PatternFill("solid", fgColor=dark_green)
    medium_green_fill = PatternFill("solid", fgColor=medium_green)
    light_green_fill = PatternFill("solid", fgColor=light_green)
    very_light_green_fill = PatternFill("solid", fgColor=very_light_green)
    orange_header_fill = PatternFill("solid", fgColor=header_orange)
    pale_yellow_fill = PatternFill("solid", fgColor=pale_yellow)
    bright_green_fill = PatternFill("solid", fgColor=bright_green)
    purple_fill = PatternFill("solid", fgColor=purple_separator)

    # Borders
    thin_grid = Side(style="thin", color="B6D7A8")
    medium_grid = Side(style="medium", color="38761D")
    dark_side = Side(style="medium", color="274E13")
    purple_side = Side(style="thick", color=purple_separator)
    red_side = Side(style="medium", color=red)
    black_side = Side(style="thin", color="000000")

    grid_border = Border(left=thin_grid, right=thin_grid, top=thin_grid, bottom=thin_grid)
    green_border = Border(left=dark_side, right=dark_side, top=dark_side, bottom=dark_side)
    purple_bottom_border = Border(bottom=purple_side)
    normal_bar_border = Border(left=black_side, right=black_side, top=black_side, bottom=black_side)
    general_bar_border = Border(left=red_side, right=red_side, top=red_side, bottom=red_side)

    # --- Τίτλος ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_time_col)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "ΠΡΟΓΡΑΜΜΑ GANTT"
    title_cell.fill = dark_green_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = green_border

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_time_col)
    subtitle_cell = ws.cell(row=2, column=1)
    subtitle_cell.value = f"Εβδομάδα: {start_of_week.strftime('%d/%m/%Y')} - {(start_of_week + timedelta(days=6)).strftime('%d/%m/%Y')}"
    subtitle_cell.fill = light_green_fill
    subtitle_cell.font = Font(color=dark_text, bold=True, size=10)
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
    subtitle_cell.border = grid_border

    # --- Αριστερές κεφαλίδες ---
    headers = {
        day_col: "ΗΜΕΡΑ",
        date_col: "ΗΜ/ΝΙΑ",
        info_col: "ΑΔΕΙΕΣ / ΔΙΑΘΕΣΙΜΑ",
        spacer_col: "ΠΡΟΓΡΑΜΜΑ",
    }

    for col, label in headers.items():
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
        cell = ws.cell(row=3, column=col)
        cell.value = label
        cell.fill = orange_header_fill if col == spacer_col else medium_green_fill
        cell.font = Font(color=dark_text, bold=True, size=8)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _apply_range_border(ws, 3, 4, col, col, green_border)

    # --- Κεφαλίδες ωρών ---
    for hour_offset in range(20):
        col_start = first_time_col + (hour_offset * (60 // slot_minutes))
        col_end = col_start + (60 // slot_minutes) - 1
        label_hour = start_hour + hour_offset
        label = f"{label_hour:02d}:00" if label_hour < 24 else "00:00"

        ws.merge_cells(start_row=3, start_column=col_start, end_row=3, end_column=col_end)
        cell = ws.cell(row=3, column=col_start)
        cell.value = label
        cell.fill = orange_header_fill
        cell.font = Font(bold=True, color=dark_text, size=8)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        _apply_range_border(ws, 3, 3, col_start, col_end, grid_border)

    for slot in range(slot_count):
        col = first_time_col + slot
        minute_from_start = slot * slot_minutes
        hour = start_hour + (minute_from_start // 60)
        minute = minute_from_start % 60
        if hour >= 24:
            hour -= 24

        cell = ws.cell(row=4, column=col)
        cell.value = f"{hour:02d}:{minute:02d}"
        cell.fill = orange_header_fill
        cell.font = Font(color=dark_text, size=7)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = grid_border

    # --- Διαστάσεις ---
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 14
    for col in range(first_time_col, last_time_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 3.7

    ws.row_dimensions[1].height = 23
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 16

    day_names_gr = ["Δευτέρα", "Τρίτη", "Τετάρτη", "Πέμπτη", "Παρασκευή", "Σάββατο", "Κυριακή"]

    current_row = 5
    for day_idx in range(7):
        current_date = start_of_week + timedelta(days=day_idx)
        day_name = day_names_gr[day_idx]

        day_groups = [
            group for group in (wk_groups or {}).values()
            if group.get("Date") == current_date
        ]

        day_info_text, day_info_lines = _day_left_label(current_date)
        placed_groups, lane_count = _build_lanes(day_groups)

        min_rows_for_label = max(1, math.ceil(day_info_lines / 2.2))
        lane_count = max(lane_count, min_rows_for_label)

        day_start_row = current_row
        day_end_row = current_row + lane_count - 1

        # A: ημέρα κάθετα
        ws.merge_cells(start_row=day_start_row, start_column=day_col, end_row=day_end_row, end_column=day_col)
        day_cell = ws.cell(row=day_start_row, column=day_col)
        day_cell.value = day_name
        day_cell.fill = dark_green_fill
        day_cell.font = Font(color="FFFFFF", bold=True, size=9)
        day_cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
        _apply_range_border(ws, day_start_row, day_end_row, day_col, day_col, green_border)

        # B: ημερομηνία
        ws.merge_cells(start_row=day_start_row, start_column=date_col, end_row=day_end_row, end_column=date_col)
        date_cell = ws.cell(row=day_start_row, column=date_col)
        date_cell.value = current_date.strftime("%d/%m/%Y")
        date_cell.fill = medium_green_fill
        date_cell.font = Font(color=dark_text, bold=True, size=8)
        date_cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
        _apply_range_border(ws, day_start_row, day_end_row, date_col, date_col, green_border)

        # C: άδειες / διαθέσιμοι
        ws.merge_cells(start_row=day_start_row, start_column=info_col, end_row=day_end_row, end_column=info_col)
        info_cell = ws.cell(row=day_start_row, column=info_col)
        info_cell.value = day_info_text
        info_cell.fill = bright_green_fill if "ΜΕΤΑ ΤΑ ΠΡΩΙΝΑ" in day_info_text else pale_yellow_fill
        info_cell.font = Font(color=dark_text, bold=True, size=7)
        info_cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        _apply_range_border(ws, day_start_row, day_end_row, info_col, info_col, green_border)

        # D: spacer/label όπως το δείγμα
        ws.merge_cells(start_row=day_start_row, start_column=spacer_col, end_row=day_end_row, end_column=spacer_col)
        spacer_cell = ws.cell(row=day_start_row, column=spacer_col)
        spacer_cell.value = ""
        spacer_cell.fill = light_green_fill
        spacer_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _apply_range_border(ws, day_start_row, day_end_row, spacer_col, spacer_col, green_border)

        # Grid φόντου
        for row in range(day_start_row, day_end_row + 1):
            ws.row_dimensions[row].height = 29
            for col in range(first_time_col, last_time_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = very_light_green_fill
                cell.border = grid_border

        # Μπάρες
        for group, lane_idx in placed_groups:
            row = day_start_row + lane_idx
            start_slot = _slot_index(group.get("StartTime"), slot_minutes=slot_minutes, mode="floor")
            end_slot = _slot_index(group.get("EndTime"), slot_minutes=slot_minutes, mode="ceil")

            start_slot = max(0, min(slot_count - 1, start_slot))
            end_slot = max(start_slot + 1, min(slot_count, end_slot))

            col_start = first_time_col + start_slot
            col_end = first_time_col + end_slot - 1

            project = str(group.get("Project", "")).upper()
            employees = ", ".join(group.get("Employees", [])).upper()
            notes = str(group.get("Notes", "") or "").upper()
            arrival = group.get("ArrivalTime", "")

            parts = []
            if arrival:
                parts.append(f"[Προσ: {arrival}]")
            parts.append(f"{group.get('StartTime')}-{group.get('EndTime')}")
            parts.append(project)
            if employees:
                parts.append(employees)
            if notes:
                parts.append(f"({notes})")

            bar_text = _shorten_text(" | ".join(parts), max_len=120)
            fill_hex = _normalize_hex(group.get("ColorHex"), default=bar_orange)

            ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
            bar_cell = ws.cell(row=row, column=col_start)
            bar_cell.value = bar_text
            bar_cell.fill = PatternFill("solid", fgColor=fill_hex)
            bar_cell.font = Font(bold=True, color=_font_color_for_fill(fill_hex), size=7)
            bar_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)

            is_general = bool(group.get("IsGeneral", False) or group.get("is_general", False))
            _apply_range_border(
                ws,
                row,
                row,
                col_start,
                col_end,
                general_bar_border if is_general else normal_bar_border,
            )

        # Μωβ διαχωριστικό ημέρας, όπως στο δείγμα.
        for col in range(1, last_time_col + 1):
            cell = ws.cell(row=day_end_row, column=col)
            cell.border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=purple_side,
            )

        current_row = day_end_row + 1

    # Πάγωμα τίτλων/headers.
    ws.freeze_panes = "E5"

    # Print / view setup
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def render_visual_gantt_excel_export(wk_groups, start_of_week):
    """Κουμπί εξαγωγής οπτικού Gantt σε Excel με συγχωνευμένα κελιά."""
    if not wk_groups:
        return

    try:
        data = create_visual_gantt_excel(wk_groups, start_of_week)
    except Exception as exc:
        st.warning(f"Δεν ήταν δυνατή η δημιουργία του οπτικού Excel Gantt: {exc}")
        return

    st.download_button(
        label="📊 Εξαγωγή Gantt σε Excel με μπάρες",
        data=data,
        file_name=f"Gantt_Bars_{start_of_week.strftime('%d_%m_%Y')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
